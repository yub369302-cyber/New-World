"""
Excel 记录模块
将每次推荐的岗位记录到 Excel 文件中，方便追踪
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.config import settings


class ExcelRecorder:
    """Excel 记录器 - 保存历史推荐岗位"""

    HEADERS = [
        "日期时间",
        "职位名称",
        "公司",
        "城市",
        "匹配分数",
        "薪资",
        "来源",
        "发布时间",
        "匹配原因",
        "简历建议",
        "招呼语",
        "链接",
    ]

    def __init__(self):
        self.filepath = settings.EXCEL_PATH

    def save_jobs(self, jobs: list[dict]) -> str:
        """
        保存岗位到 Excel

        Args:
            jobs: 分析后的岗位列表

        Returns:
            Excel 文件路径
        """
        if not jobs:
            return ""

        # 加载或创建工作簿
        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "岗位推荐记录"
            self._write_headers(ws)

        # 写入数据
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        for job in jobs:
            row = [
                now,
                job.get("title", ""),
                job.get("company", ""),
                job.get("location", ""),
                job.get("match_score", 0),
                job.get("salary", ""),
                job.get("source", ""),
                job.get("posted_at", ""),
                job.get("match_reason", ""),
                job.get("resume_advice", ""),
                job.get("greeting", ""),
                job.get("link", ""),
            ]
            ws.append(row)

        # 调整列宽
        self._auto_width(ws)

        # 保存
        wb.save(self.filepath)
        print(f"[成功] 已保存 {len(jobs)} 条记录到 {self.filepath}")
        return self.filepath

    def _write_headers(self, ws):
        """写入表头并设置样式"""
        ws.append(self.HEADERS)

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, _ in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _auto_width(self, ws):
        """自动调整列宽"""
        column_widths = [18, 25, 20, 10, 8, 12, 15, 12, 30, 40, 40, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = width

    def get_history(self, limit: int = 50) -> list[dict]:
        """
        获取历史推荐记录

        Args:
            limit: 返回条数

        Returns:
            历史记录列表
        """
        if not os.path.exists(self.filepath):
            return []

        wb = load_workbook(self.filepath)
        ws = wb.active
        records = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 有日期数据
                records.append({
                    "date": row[0],
                    "title": row[1],
                    "company": row[2],
                    "location": row[3],
                    "score": row[4],
                    "salary": row[5],
                    "source": row[6],
                })

        # 返回最近的记录
        return records[-limit:]
